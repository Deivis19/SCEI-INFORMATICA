import os
import sys
from pathlib import Path
from PyQt6.QtWidgets import QApplication, QTableWidget, QMessageBox, QTableWidgetItem
from PyQt6.QtGui import QIcon, QPixmap, QPainter, QPainterPath, QColor
from PyQt6.QtCore import Qt

class NumericTableWidgetItem(QTableWidgetItem):
    def __lt__(self, other):
        try:
            return int(self.text()) < int(other.text())
        except ValueError:
            return self.text() < other.text()

def resource_path(*parts: str) -> str:
    # Resolver ruta de recursos compatible con entorno congelado (PyInstaller)
    # Asumimos que este archivo está en scei/utils.py
    base_dir = os.path.dirname(os.path.abspath(__file__))
    try:
        if getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS"):
            base_dir = sys._MEIPASS  # type: ignore[attr-defined]
    except Exception:
        pass
    
    # Si estamos en scei/, recursos está en scei/resources o en root/resources?
    # En main.py original: os.path.join(BASE_DIR, "resources", *parts) donde BASE_DIR era dirname(abspath(__file__))
    # Aquí base_dir es scei/
    return os.path.join(base_dir, "resources", *parts)

def load_icon(name: str) -> QIcon:
    path = resource_path("icons", name)
    if os.path.exists(path):
        return QIcon(path)
    return QIcon()

def load_pixmap(name: str) -> QPixmap:
    path = resource_path("images", name)
    if os.path.exists(path):
        return QPixmap(path)
    return QPixmap()

def circular_pixmap(pixmap: QPixmap, size: int) -> QPixmap:
    if pixmap.isNull() or size <= 0:
        return QPixmap()
    target = QPixmap(size, size)
    target.fill(Qt.GlobalColor.transparent)
    painter = QPainter(target)
    painter.setRenderHint(QPainter.RenderHint.Antialiasing)
    path = QPainterPath()
    path.addEllipse(0, 0, size, size)
    painter.setClipPath(path)
    scaled = pixmap.scaled(size, size, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation)
    x = int((size - scaled.width()) / 2)
    y = int((size - scaled.height()) / 2)
    painter.drawPixmap(x, y, scaled)
    painter.end()
    return target

def apply_light_theme(app: QApplication) -> None:
    theme_path = resource_path("theme", "light.qss")
    if os.path.exists(theme_path):
        with open(theme_path, "r", encoding="utf-8") as f:
            app.setStyleSheet(f.read())

def export_table_to_excel(table: QTableWidget, filename: str) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except Exception:
        QMessageBox.warning(table, "Exportar", "No se encontró openpyxl. Instale dependencias.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"

    headers = [table.horizontalHeaderItem(c).text() for c in range(table.columnCount())]
    ws.append(headers)

    for r in range(table.rowCount()):
        row = []
        for c in range(table.columnCount()):
            item = table.item(r, c)
            row.append(item.text() if item else "")
        ws.append(row)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="34495E")
    data_alignment = Alignment(vertical="center", wrap_text=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    border_style = Border(bottom=Side(style="thin", color="CBD5E1"))

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border_style

    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = data_alignment
            cell.border = border_style
        if row_idx % 2 == 0:
            stripe_fill = PatternFill(fill_type="solid", fgColor="F8FAFC")
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = stripe_fill

    for column_cells in ws.columns:
        max_length = 12
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            try:
                cell_length = len(str(cell.value)) if cell.value is not None else 0
                if cell_length > max_length:
                    max_length = cell_length
            except Exception:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 4, 50)

    if ws.max_row > 1:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    ws.row_dimensions[1].height = 22
    wb.save(filename)

def validate_password_strength(password: str) -> bool:
    """Valida que la contraseña tenga mayúscula, número y caracter especial."""
    has_upper = any(c.isupper() for c in password)
    has_digit = any(c.isdigit() for c in password)
    has_special = any(not c.isalnum() for c in password)
    return has_upper and has_digit and has_special
